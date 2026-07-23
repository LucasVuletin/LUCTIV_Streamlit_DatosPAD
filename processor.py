from __future__ import annotations

from copy import copy
from dataclasses import dataclass, field
from io import BytesIO
import math
from pathlib import Path
import re
import unicodedata
from typing import BinaryIO, Iterable

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


REQUIRED_SHEETS = ("Input", "Survey", "Punzados")
TEMPLATE_SHEET = "Datos terminados"
ALLOWED_EXTENSIONS = {".xlsm", ".xlsx"}
FORMULA_ERROR_VALUES = {"#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A"}
GENERAL_NUMBER_FORMAT = "General"
INTEGER_NUMBER_FORMAT = "0"


class LuctivError(Exception):
    """Base error shown to the end user."""


class InvalidWorkbookError(LuctivError):
    """Raised when the uploaded workbook does not match the expected structure."""


@dataclass(frozen=True)
class FractureConfig:
    label: str
    start_stage: int
    end_stage: int
    clusters: int
    spf: int


@dataclass(frozen=True)
class SurveyPoint:
    md: float
    inclination: float
    azimuth: float
    tvd: float


@dataclass(frozen=True)
class Cluster:
    cluster_number: int
    top_md: float
    base_md: float
    stage: int
    spf: int
    expected_clusters: int | None = None
    stage_length: float | None = None
    override_note: str | None = None


@dataclass(frozen=True)
class StageInterval:
    stage: int
    top_md: float
    base_md: float
    plug_md: float


@dataclass
class ProcessingResult:
    well_name: str
    output_filename: str
    fracture_configs: list[FractureConfig]
    survey: list[SurveyPoint]
    clusters: list[Cluster]
    stages: list[StageInterval]
    warnings: list[str] = field(default_factory=list)
    checks: list[str] = field(default_factory=list)

    @property
    def wellbore_row_count(self) -> int:
        return len(self.stages) * 2

    @property
    def override_count(self) -> int:
        return sum(1 for cluster in self.clusters if cluster.override_note)


@dataclass(frozen=True)
class GeneratedWorkbook:
    result: ProcessingResult
    data: bytes


def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text)


def _is_number(value: object) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def _coerce_float(value: object) -> float | None:
    if _is_number(value):
        return float(value)
    if isinstance(value, str):
        cleaned = value.strip().replace(" ", "")
        if not cleaned:
            return None
        if "," in cleaned and "." in cleaned:
            if cleaned.rfind(",") > cleaned.rfind("."):
                cleaned = cleaned.replace(".", "").replace(",", ".")
            else:
                cleaned = cleaned.replace(",", "")
        else:
            cleaned = cleaned.replace(",", ".")
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def _as_float(value: object, field_name: str, row_number: int) -> float:
    number = _coerce_float(value)
    if number is not None:
        return number
    raise InvalidWorkbookError(
        f"El campo '{field_name}' de la fila {row_number} no es numérico: {value!r}."
    )


def _as_int(value: object, field_name: str, row_number: int) -> int:
    number = _as_float(value, field_name, row_number)
    if not number.is_integer():
        raise InvalidWorkbookError(
            f"El campo '{field_name}' de la fila {row_number} debe ser entero: {value!r}."
        )
    return int(number)


def _find_header_row(
    sheet: Worksheet,
    required_terms: Iterable[str],
    max_scan_rows: int = 60,
) -> int:
    required = [_normalize_text(term) for term in required_terms]
    for row_idx in range(1, min(sheet.max_row, max_scan_rows) + 1):
        values = [_normalize_text(cell.value) for cell in sheet[row_idx]]
        if all(any(term == value or term in value for value in values) for term in required):
            return row_idx
    raise InvalidWorkbookError(
        f"No se pudo localizar el encabezado esperado en la hoja '{sheet.title}'."
    )


def _find_column(headers: dict[int, str], aliases: Iterable[str]) -> int:
    normalized_aliases = [_normalize_text(alias) for alias in aliases]
    for alias in normalized_aliases:
        for col_idx, header in headers.items():
            if header == alias:
                return col_idx
    for alias in normalized_aliases:
        for col_idx, header in headers.items():
            if alias in header:
                return col_idx
    raise InvalidWorkbookError(
        f"Falta una columna requerida. Se buscó: {', '.join(aliases)}."
    )


def _find_column_or_none(headers: dict[int, str], aliases: Iterable[str]) -> int | None:
    try:
        return _find_column(headers, aliases)
    except InvalidWorkbookError:
        return None


def _extract_well_name(input_sheet: Worksheet, fallback_filename: str) -> str:
    for row in input_sheet.iter_rows(min_row=1, max_row=min(input_sheet.max_row, 20)):
        for idx, cell in enumerate(row):
            if _normalize_text(cell.value).rstrip(":") == "nombre":
                if idx + 1 < len(row) and row[idx + 1].value:
                    return str(row[idx + 1].value).strip()
    return Path(fallback_filename).stem.replace("_version2", "")


def _safe_output_filename(well_name: str) -> str:
    cleaned = well_name.replace("(h)", "h").replace("(H)", "H")
    cleaned = re.sub(r"[()]", "", cleaned)
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "-", cleaned).strip("-_.")
    return f"{cleaned or 'pozo'}_datos_terminados.xlsx"


def _extract_fracture_configs(sheet: Worksheet) -> list[FractureConfig]:
    header_row = _find_header_row(sheet, ("Etapas", "Inicio", "Fin", "SPF"))
    headers = {
        cell.column: _normalize_text(cell.value)
        for cell in sheet[header_row]
        if cell.value is not None
    }
    col_label = _find_column(headers, ("Etapas",))
    col_start = _find_column(headers, ("Inicio",))
    col_end = _find_column(headers, ("Fin",))
    col_clusters = _find_column(
        headers,
        ("N° Cl", "N Cl", "Nro Cl", "Cantidad de clusters", "Clusters"),
    )
    col_spf = _find_column(headers, ("SPF",))

    configs: list[FractureConfig] = []
    blank_run = 0
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        label = sheet.cell(row_idx, col_label).value
        start = sheet.cell(row_idx, col_start).value
        end = sheet.cell(row_idx, col_end).value

        if label in (None, "") and start in (None, "") and end in (None, ""):
            blank_run += 1
            if configs and blank_run >= 3:
                break
            continue
        blank_run = 0

        if not (_is_number(start) or isinstance(start, str)):
            continue
        if not (_is_number(end) or isinstance(end, str)):
            continue

        configs.append(
            FractureConfig(
                label=str(label).strip() if label not in (None, "") else f"Etapas {start}-{end}",
                start_stage=_as_int(start, "Inicio", row_idx),
                end_stage=_as_int(end, "Fin", row_idx),
                clusters=_as_int(sheet.cell(row_idx, col_clusters).value, "N° Cl", row_idx),
                spf=_as_int(sheet.cell(row_idx, col_spf).value, "SPF", row_idx),
            )
        )

    if not configs:
        raise InvalidWorkbookError("No se encontraron configuraciones de fractura en 'Input'.")
    return configs


def _extract_survey(sheet: Worksheet) -> list[SurveyPoint]:
    header_row = _find_header_row(sheet, ("MD", "TVD", "INCL"), max_scan_rows=15)
    headers = {
        cell.column: _normalize_text(cell.value)
        for cell in sheet[header_row]
        if cell.value is not None
    }
    col_md = _find_column(headers, ("MD",))
    col_tvd = _find_column(headers, ("TVD",))
    col_incl = _find_column(headers, ("INCL", "Inclination"))
    col_azim = _find_column(headers, ("AZIM_TN", "Azimuth", "AZIM"))

    survey: list[SurveyPoint] = []
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        md = sheet.cell(row_idx, col_md).value
        md_number = _coerce_float(md)
        if md_number is None:
            continue
        survey.append(
            SurveyPoint(
                md=round(md_number, 2),
                inclination=round(_as_float(sheet.cell(row_idx, col_incl).value, "INCL", row_idx), 2),
                azimuth=round(_as_float(sheet.cell(row_idx, col_azim).value, "AZIM", row_idx), 2),
                tvd=round(_as_float(sheet.cell(row_idx, col_tvd).value, "TVD", row_idx), 2),
            )
        )

    if not survey:
        raise InvalidWorkbookError("No se encontraron registros numéricos en la hoja 'Survey'.")
    return survey


def _extract_clusters(sheet: Worksheet) -> list[Cluster]:
    header_row = _find_header_row(
        sheet,
        ("# Cluster", "Tope Cluster MD", "Base Cluster MD", "Número etapa"),
        max_scan_rows=15,
    )
    headers = {
        cell.column: _normalize_text(cell.value)
        for cell in sheet[header_row]
        if cell.value is not None
    }
    col_number = _find_column(headers, ("# Cluster", "Cluster"))
    col_top = _find_column(headers, ("Tope Cluster MD",))
    col_base = _find_column(headers, ("Base Cluster MD",))
    col_stage = _find_column(headers, ("Número etapa", "Numero etapa"))
    col_spf = _find_column(headers, ("N° de tiros x cluster", "tiros x cluster"))
    col_stage_length = _find_column_or_none(headers, ("Longitud de etapa",))
    col_expected = _find_column_or_none(headers, ("Cantidad de clusters",))

    override_col: int | None = None
    for col_idx, header in headers.items():
        if "sobreescribir" in header or "cambio de punzados" in header:
            override_col = col_idx
            break

    clusters: list[Cluster] = []
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        cluster_number = sheet.cell(row_idx, col_number).value
        stage = sheet.cell(row_idx, col_stage).value
        cluster_number_value = _coerce_float(cluster_number)
        stage_value = _coerce_float(stage)
        if cluster_number_value is None or stage_value is None:
            continue

        override_value = sheet.cell(row_idx, override_col).value if override_col else None
        expected_value = sheet.cell(row_idx, col_expected).value if col_expected else None
        stage_length_value = sheet.cell(row_idx, col_stage_length).value if col_stage_length else None
        clusters.append(
            Cluster(
                cluster_number=_as_int(cluster_number_value, "# Cluster", row_idx),
                top_md=_as_float(sheet.cell(row_idx, col_top).value, "Tope Cluster MD", row_idx),
                base_md=_as_float(sheet.cell(row_idx, col_base).value, "Base Cluster MD", row_idx),
                stage=_as_int(stage_value, "Número etapa", row_idx),
                spf=_as_int(sheet.cell(row_idx, col_spf).value, "N° de tiros x cluster", row_idx),
                expected_clusters=(
                    _as_int(expected_value, "Cantidad de clusters", row_idx)
                    if expected_value not in (None, "")
                    else None
                ),
                stage_length=(
                    _as_float(stage_length_value, "Longitud de etapa", row_idx)
                    if stage_length_value not in (None, "")
                    else None
                ),
                override_note=(str(override_value).strip() if override_value not in (None, "") else None),
            )
        )

    if not clusters:
        raise InvalidWorkbookError("No se encontraron clústeres numéricos en la hoja 'Punzados'.")
    return clusters


def _build_stages(
    clusters: list[Cluster],
    configs: list[FractureConfig],
) -> tuple[list[StageInterval], list[str], list[str]]:
    grouped: dict[int, list[Cluster]] = {}
    for cluster in clusters:
        grouped.setdefault(cluster.stage, []).append(cluster)

    stages: list[StageInterval] = []
    warnings: list[str] = []
    checks: list[str] = []

    actual_stage_numbers = sorted(grouped)
    expected_stage_numbers = list(range(1, max(actual_stage_numbers) + 1))
    if actual_stage_numbers != expected_stage_numbers:
        missing = sorted(set(expected_stage_numbers) - set(actual_stage_numbers))
        raise InvalidWorkbookError(
            "La numeración de etapas no es consecutiva. "
            f"Etapas faltantes: {missing or 'no identificadas'}."
        )

    for cluster in clusters:
        if cluster.top_md >= cluster.base_md:
            raise InvalidWorkbookError(
                f"Clúster {cluster.cluster_number}, etapa {cluster.stage}: "
                f"el Tope ({cluster.top_md}) debe ser menor que el Fondo ({cluster.base_md})."
            )

    seen_cluster_numbers: set[int] = set()
    duplicates: list[int] = []
    for cluster in clusters:
        if cluster.cluster_number in seen_cluster_numbers:
            duplicates.append(cluster.cluster_number)
        seen_cluster_numbers.add(cluster.cluster_number)
    if duplicates:
        raise InvalidWorkbookError(
            f"Hay números de clúster duplicados: {sorted(set(duplicates))[:20]}."
        )

    for stage_number in actual_stage_numbers:
        stage_clusters = grouped[stage_number]
        top_md = min(cluster.top_md for cluster in stage_clusters)
        base_md = max(cluster.base_md for cluster in stage_clusters)
        expected_counts = {
            cluster.expected_clusters
            for cluster in stage_clusters
            if cluster.expected_clusters is not None
        }
        spfs = {cluster.spf for cluster in stage_clusters}
        stage_lengths = {
            round(cluster.stage_length, 6)
            for cluster in stage_clusters
            if cluster.stage_length is not None
        }

        if len(expected_counts) > 1:
            raise InvalidWorkbookError(
                f"Etapa {stage_number}: la cantidad esperada de clústeres no es consistente."
            )
        if expected_counts and len(stage_clusters) != next(iter(expected_counts)):
            raise InvalidWorkbookError(
                f"Etapa {stage_number}: se encontraron {len(stage_clusters)} clústeres "
                f"y se esperaban {next(iter(expected_counts))} según Punzados."
            )
        if len(spfs) != 1:
            raise InvalidWorkbookError(f"Etapa {stage_number}: hay valores SPF inconsistentes.")
        if len(stage_lengths) != 1:
            warnings.append(
                f"Etapa {stage_number}: aparecen longitudes de etapa diferentes dentro del mismo grupo."
            )
        if top_md >= base_md:
            raise InvalidWorkbookError(
                f"Etapa {stage_number}: el Tope ({top_md}) debe ser menor que el Fondo ({base_md})."
            )

        stages.append(
            StageInterval(
                stage=stage_number,
                top_md=round(top_md, 2),
                base_md=round(base_md, 2),
                plug_md=round(base_md + 3.7, 2),
            )
        )

    # Cross-check every configured range against the actual clusters and SPF.
    configured_stages: set[int] = set()
    for config in configs:
        if config.start_stage > config.end_stage:
            raise InvalidWorkbookError(
                f"Configuración '{config.label}': Inicio es mayor que Fin."
            )
        for stage_number in range(config.start_stage, config.end_stage + 1):
            if stage_number in configured_stages:
                raise InvalidWorkbookError(
                    f"La etapa {stage_number} aparece en más de una configuración de fractura."
                )
            configured_stages.add(stage_number)
            stage_clusters = grouped.get(stage_number)
            if not stage_clusters:
                raise InvalidWorkbookError(
                    f"La configuración '{config.label}' incluye la etapa {stage_number}, "
                    "pero no hay punzados para esa etapa."
                )
            if len(stage_clusters) != config.clusters:
                raise InvalidWorkbookError(
                    f"{config.label}, etapa {stage_number}: se encontraron "
                    f"{len(stage_clusters)} clústeres y la configuración indica {config.clusters}."
                )
            actual_spf = {cluster.spf for cluster in stage_clusters}
            if actual_spf != {config.spf}:
                raise InvalidWorkbookError(
                    f"{config.label}, etapa {stage_number}: SPF {sorted(actual_spf)} "
                    f"y la configuración indica {config.spf}."
                )

    unconfigured = sorted(set(actual_stage_numbers) - configured_stages)
    if unconfigured:
        raise InvalidWorkbookError(
            f"Hay etapas sin configuración de fractura: {unconfigured}."
        )

    overrides = [cluster for cluster in clusters if cluster.override_note]
    if overrides:
        warnings.append(
            f"Se detectaron {len(overrides)} celdas con observaciones o sobreescrituras "
            "en la hoja Punzados. Revisar el archivo original si el cambio modifica Tope o Fondo."
        )

    checks.extend(
        [
            f"{len(stages)} etapas consecutivas",
            f"{len(clusters)} clústeres procesados",
            "Cantidad de clústeres por etapa correcta",
            "SPF consistente con las configuraciones",
            f"{len(configs)} configuraciones de fractura",
        ]
    )
    return stages, warnings, checks


def analyze_workbook(file_obj: bytes | BinaryIO, filename: str) -> ProcessingResult:
    stream = BytesIO(file_obj) if isinstance(file_obj, bytes) else file_obj
    try:
        workbook = load_workbook(
            stream,
            read_only=False,
            data_only=True,
            keep_vba=False,
            keep_links=False,
        )
    except Exception as exc:  # openpyxl exposes several parser exceptions
        raise InvalidWorkbookError(
            "No se pudo abrir el archivo. Verificá que sea un Excel .xlsm o .xlsx válido."
        ) from exc

    missing = [name for name in REQUIRED_SHEETS if name not in workbook.sheetnames]
    if missing:
        quoted = ", ".join(f'"{name}"' for name in missing)
        if len(missing) == 1:
            raise InvalidWorkbookError(
                f"No se encontró la hoja {quoted}. "
                "Verificá que estés cargando el archivo Version 2 correspondiente al pozo."
            )
        raise InvalidWorkbookError(
            f"No se encontraron las hojas {quoted}. "
            "Verificá que estés cargando el archivo Version 2 correspondiente al pozo."
        )

    input_sheet = workbook["Input"]
    well_name = _extract_well_name(input_sheet, filename)
    configs = _extract_fracture_configs(input_sheet)
    survey = _extract_survey(workbook["Survey"])
    clusters = _extract_clusters(workbook["Punzados"])
    stages, warnings, checks = _build_stages(clusters, configs)
    checks.extend(
        [
            f"{len(survey)} registros Survey",
            f"{len(stages) * 2} filas Wellbore IFS",
            "Tapones calculados como Fondo + 3,7 m",
        ]
    )

    return ProcessingResult(
        well_name=well_name,
        output_filename=_safe_output_filename(well_name),
        fracture_configs=configs,
        survey=survey,
        clusters=clusters,
        stages=stages,
        warnings=warnings,
        checks=checks,
    )


def _check_generated_workbook(data: bytes, result: ProcessingResult) -> list[str]:
    try:
        workbook = load_workbook(BytesIO(data), data_only=False)
    except Exception as exc:
        raise LuctivError("El archivo generado no pudo volver a abrirse con openpyxl.") from exc

    if TEMPLATE_SHEET not in workbook.sheetnames:
        raise LuctivError(f"El archivo generado no contiene la hoja '{TEMPLATE_SHEET}'.")

    sheet = workbook[TEMPLATE_SHEET]
    errors: list[str] = []
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip() in FORMULA_ERROR_VALUES:
                errors.append(f"{cell.coordinate}: {cell.value.strip()}")
    if errors:
        raise LuctivError(
            "El archivo generado contiene errores visibles de Excel: " + ", ".join(errors[:20])
        )

    for row_idx, stage in enumerate(result.stages, start=4):
        plug = sheet.cell(row_idx, 16).value
        fondo = sheet.cell(row_idx, 15).value
        if not isinstance(plug, (int, float)) or not isinstance(fondo, (int, float)):
            raise LuctivError(f"El tapón de la fila {row_idx} no quedó como valor numérico.")
        if not math.isclose(float(plug), float(fondo) + 3.7, abs_tol=0.001):
            raise LuctivError(
                f"El tapón de la etapa {stage.stage} no coincide con Fondo + 3,7 m."
            )

    expected_wellbore: list[tuple[str, float, float]] = []
    for stage in reversed(result.stages):
        expected_wellbore.append(("Treatment Interval", stage.top_md, stage.base_md))
        expected_wellbore.append(("Perforations", stage.top_md, stage.base_md))

    for offset, expected in enumerate(expected_wellbore):
        row_idx = 4 + offset
        label, top_md, base_md = expected
        if sheet.cell(row_idx, 18).value != label:
            raise LuctivError(f"Wellbore IFS no tiene la etiqueta esperada en la fila {row_idx}.")
        if not math.isclose(float(sheet.cell(row_idx, 19).value), top_md, abs_tol=0.001):
            raise LuctivError(f"Wellbore IFS tiene un Tope incorrecto en la fila {row_idx}.")
        if not math.isclose(float(sheet.cell(row_idx, 20).value), base_md, abs_tol=0.001):
            raise LuctivError(f"Wellbore IFS tiene un Fondo incorrecto en la fila {row_idx}.")

    variable_blocks = (
        ("Datos Fractura", 3 + len(result.fracture_configs), 1, 5),
        ("Datos Survey", 4 + len(result.survey), 8, 11),
        ("Smart Staging", 4 + len(result.stages), 13, 16),
        ("Wellbore IFS", 4 + result.wellbore_row_count, 18, 20),
    )
    for block_name, first_empty_row, min_col, max_col in variable_blocks:
        for row_idx in range(first_empty_row, sheet.max_row + 1):
            for col_idx in range(min_col, max_col + 1):
                value = sheet.cell(row_idx, col_idx).value
                if value not in (None, ""):
                    raise LuctivError(
                        f"Quedaron datos residuales en {block_name}: "
                        f"{sheet.cell(row_idx, col_idx).coordinate}."
                    )

    return [
        "Wellbore IFS verificado con dos filas por etapa",
        "Archivo .xlsx verificado con openpyxl",
        "Sin datos residuales en los rangos variables",
        "Sin errores críticos",
    ]


def _copy_cell_style(source: Cell, target: Cell) -> None:
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.alignment:
        target.alignment = copy(source.alignment)
    if source.protection:
        target.protection = copy(source.protection)


def _copy_row_style(
    sheet: Worksheet,
    source_row: int,
    target_row: int,
    min_col: int,
    max_col: int,
) -> None:
    for col_idx in range(min_col, max_col + 1):
        _copy_cell_style(sheet.cell(source_row, col_idx), sheet.cell(target_row, col_idx))
    if source_row in sheet.row_dimensions:
        sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height


def _clear_values(sheet: Worksheet, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in sheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
    ):
        for cell in row:
            cell.value = None


def _ensure_styles(sheet: Worksheet, result: ProcessingResult) -> None:
    design_end = 2 + len(result.fracture_configs)
    survey_end = 3 + len(result.survey)
    stage_end = 3 + len(result.stages)
    wellbore_end = 3 + result.wellbore_row_count

    # The template has representative styles in these rows. Copy only styling, never values.
    for row_idx in range(3, design_end + 1):
        source_row = 3 if row_idx == 3 else 4
        _copy_row_style(sheet, source_row, row_idx, 1, 5)

    for row_idx in range(4, survey_end + 1):
        _copy_row_style(sheet, 4, row_idx, 8, 11)

    for row_idx in range(4, stage_end + 1):
        _copy_row_style(sheet, 4, row_idx, 13, 16)

    for row_idx in range(4, wellbore_end + 1):
        source_row = 4 if (row_idx - 4) % 2 == 0 else 5
        _copy_row_style(sheet, source_row, row_idx, 18, 20)


def _set_number_format(
    sheet: Worksheet,
    min_row: int,
    max_row: int,
    columns: Iterable[int],
    number_format: str,
) -> None:
    if max_row < min_row:
        return
    for row_idx in range(min_row, max_row + 1):
        for col_idx in columns:
            sheet.cell(row_idx, col_idx).number_format = number_format


def _normalize_output_number_formats(sheet: Worksheet, result: ProcessingResult) -> None:
    _set_number_format(
        sheet,
        3,
        2 + len(result.fracture_configs),
        range(2, 6),
        INTEGER_NUMBER_FORMAT,
    )
    _set_number_format(
        sheet,
        4,
        3 + len(result.survey),
        range(8, 12),
        GENERAL_NUMBER_FORMAT,
    )
    _set_number_format(sheet, 4, 3 + len(result.stages), (13,), INTEGER_NUMBER_FORMAT)
    _set_number_format(
        sheet,
        4,
        3 + len(result.stages),
        range(14, 17),
        GENERAL_NUMBER_FORMAT,
    )
    _set_number_format(
        sheet,
        4,
        3 + result.wellbore_row_count,
        range(19, 21),
        GENERAL_NUMBER_FORMAT,
    )


def generate_finished_workbook(
    result: ProcessingResult,
    template_path: str | Path,
) -> GeneratedWorkbook:
    template_path = Path(template_path)
    if not template_path.exists():
        raise LuctivError("No se encontró la plantilla de salida de LUCTIV.")

    workbook = load_workbook(template_path)
    if TEMPLATE_SHEET not in workbook.sheetnames:
        raise LuctivError(
            f"La plantilla no contiene la hoja '{TEMPLATE_SHEET}'."
        )
    sheet = workbook[TEMPLATE_SHEET]

    required_max_row = max(
        3 + len(result.survey),
        3 + len(result.stages),
        3 + result.wellbore_row_count,
        2 + len(result.fracture_configs),
        sheet.max_row,
    )
    _clear_values(sheet, 3, required_max_row, 1, 5)
    _clear_values(sheet, 4, required_max_row, 8, 11)
    _clear_values(sheet, 4, required_max_row, 13, 16)
    _clear_values(sheet, 4, required_max_row, 18, 20)
    _ensure_styles(sheet, result)

    for row_idx, config in enumerate(result.fracture_configs, start=3):
        values = (
            config.label,
            config.start_stage,
            config.end_stage,
            config.clusters,
            config.spf,
        )
        for col_idx, value in enumerate(values, start=1):
            sheet.cell(row_idx, col_idx, value)

    for row_idx, point in enumerate(result.survey, start=4):
        sheet.cell(row_idx, 8, point.md)
        sheet.cell(row_idx, 9, point.inclination)
        sheet.cell(row_idx, 10, point.azimuth)
        sheet.cell(row_idx, 11, point.tvd)

    for row_idx, stage in enumerate(result.stages, start=4):
        sheet.cell(row_idx, 13, stage.stage)
        sheet.cell(row_idx, 14, stage.top_md)
        sheet.cell(row_idx, 15, stage.base_md)
        sheet.cell(row_idx, 16, stage.plug_md)

    wellbore_row = 4
    for stage in reversed(result.stages):
        for label in ("Treatment Interval", "Perforations"):
            sheet.cell(wellbore_row, 18, label)
            sheet.cell(wellbore_row, 19, stage.top_md)
            sheet.cell(wellbore_row, 20, stage.base_md)
            wellbore_row += 1

    _normalize_output_number_formats(sheet, result)

    # Ensure Excel recalculates any formulas that could exist in the template.
    try:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
    except AttributeError:
        pass

    output = BytesIO()
    workbook.save(output)
    data = output.getvalue()
    for check in _check_generated_workbook(data, result):
        if check not in result.checks:
            result.checks.append(check)
    return GeneratedWorkbook(result=result, data=data)


def process_uploaded_workbook(
    file_bytes: bytes,
    filename: str,
    template_path: str | Path,
) -> GeneratedWorkbook:
    if Path(filename).suffix.lower() not in ALLOWED_EXTENSIONS:
        raise InvalidWorkbookError("LUCTIV solo acepta archivos Excel .xlsm o .xlsx.")
    result = analyze_workbook(file_bytes, filename)
    return generate_finished_workbook(result, template_path)
