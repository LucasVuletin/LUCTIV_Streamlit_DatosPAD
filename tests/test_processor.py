from __future__ import annotations

from io import BytesIO
import os
from pathlib import Path

from openpyxl import Workbook, load_workbook
import pytest

from processor import (
    InvalidWorkbookError,
    analyze_workbook,
    generate_finished_workbook,
    process_uploaded_workbook,
)


ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "assets" / "plantilla_datos_terminados.xlsx"
SAMPLES = Path(os.environ.get("LUCTIV_SAMPLE_DIR", ROOT / "samples"))


def _xlsx_bytes(workbook: Workbook) -> bytes:
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _default_configs(stage_count: int, clusters: int = 10, spf: int = 4) -> list[dict[str, int | str]]:
    return [
        {
            "label": f"Etapas 1-{stage_count}",
            "start": 1,
            "end": stage_count,
            "clusters": clusters,
            "spf": spf,
        }
    ]


def _config_for_stage(configs: list[dict[str, int | str]], stage: int) -> dict[str, int | str]:
    for config in configs:
        if int(config["start"]) <= stage <= int(config["end"]):
            return config
    return {"label": "sin config", "start": stage, "end": stage, "clusters": 10, "spf": 4}


def make_source_workbook(
    *,
    well_name: str = "LajE-32(h)",
    stage_count: int = 3,
    survey_rows: int = 5,
    configs: list[dict[str, int | str]] | None = None,
    stage_numbers: list[int] | None = None,
    clusters_by_stage: dict[int, int] | None = None,
    spf_by_stage: dict[int, int] | None = None,
    duplicate_cluster_number: bool = False,
    bad_depth_stage: int | None = None,
    overrides: dict[tuple[int, int], str] | None = None,
    omit_sheets: set[str] | None = None,
    empty_punzados: bool = False,
) -> bytes:
    input_configs = _default_configs(stage_count) if configs is None else configs
    cluster_configs = input_configs or _default_configs(stage_count)
    stage_numbers = stage_numbers or list(range(1, stage_count + 1))
    clusters_by_stage = clusters_by_stage or {}
    spf_by_stage = spf_by_stage or {}
    overrides = overrides or {}
    omit_sheets = omit_sheets or set()

    workbook = Workbook()
    input_sheet = workbook.active
    input_sheet.title = "Input"
    survey_sheet = workbook.create_sheet("Survey")
    punzados_sheet = workbook.create_sheet("Punzados")

    input_sheet["A1"] = "Nombre"
    input_sheet["B1"] = well_name
    input_sheet.append([])
    input_sheet.append(["Etapas", "Inicio", "Fin", "N° Cl", "SPF"])
    for config in input_configs:
        input_sheet.append(
            [
                config["label"],
                config["start"],
                config["end"],
                config["clusters"],
                config["spf"],
            ]
        )

    survey_sheet.append(["fila separadora", None, None, None])
    survey_sheet.append(["MD", "INCL", "AZIM_TN", "TVD"])
    survey_sheet.append(["texto", "sin", "md", "numerico"])
    for idx in range(survey_rows):
        survey_sheet.append(
            [
                f"{1000 + idx * 10:.2f}" if idx == 0 else 1000 + idx * 10,
                80 + idx * 0.1,
                120 + idx * 0.2,
                900 + idx * 8.5,
            ]
        )

    punzados_sheet.append(
        [
            "# Cluster",
            "Tope Cluster MD (m)",
            "Base Cluster MD (m)",
            "Número etapa",
            "N° de tiros x cluster",
            "Longitud de etapa",
            "Cantidad de clusters",
            "En caso de cambio de punzados, sobreescribir datos (NO BORRAR)",
        ]
    )
    cluster_number = 1
    if not empty_punzados:
        for stage in stage_numbers:
            config = _config_for_stage(cluster_configs, stage)
            cluster_count = clusters_by_stage.get(stage, int(config["clusters"]))
            spf = spf_by_stage.get(stage, int(config["spf"]))
            for index in range(cluster_count):
                top_md = 6000 + stage * 20 + index * 1.5
                base_md = top_md + 1.1
                if bad_depth_stage == stage:
                    base_md = top_md - 0.5
                number = 1 if duplicate_cluster_number and cluster_number == 2 else cluster_number
                punzados_sheet.append(
                    [
                        number,
                        top_md,
                        base_md,
                        stage,
                        spf,
                        50,
                        int(config["clusters"]),
                        overrides.get((stage, index + 1)),
                    ]
                )
                cluster_number += 1

    for sheet_name in omit_sheets:
        del workbook[sheet_name]

    return _xlsx_bytes(workbook)


def test_missing_required_sheet_names_the_exact_sheet():
    data = make_source_workbook(omit_sheets={"Punzados"})

    with pytest.raises(InvalidWorkbookError, match='No se encontró la hoja "Punzados"'):
        analyze_workbook(data, "pozo.xlsx")


def test_empty_survey_is_rejected():
    data = make_source_workbook(survey_rows=0)

    with pytest.raises(InvalidWorkbookError, match="Survey"):
        analyze_workbook(data, "pozo.xlsx")


def test_empty_punzados_is_rejected():
    data = make_source_workbook(empty_punzados=True)

    with pytest.raises(InvalidWorkbookError, match="Punzados"):
        analyze_workbook(data, "pozo.xlsx")


def test_non_consecutive_stages_are_rejected():
    data = make_source_workbook(stage_count=3, stage_numbers=[1, 3])

    with pytest.raises(InvalidWorkbookError, match="Etapas faltantes: \\[2\\]"):
        analyze_workbook(data, "pozo.xlsx")


def test_wrong_cluster_count_is_rejected():
    data = make_source_workbook(stage_count=2, clusters_by_stage={2: 8})

    with pytest.raises(InvalidWorkbookError, match="se encontraron 8 clústeres"):
        analyze_workbook(data, "pozo.xlsx")


def test_wrong_spf_is_rejected():
    data = make_source_workbook(stage_count=2, spf_by_stage={2: 5})

    with pytest.raises(InvalidWorkbookError, match="SPF"):
        analyze_workbook(data, "pozo.xlsx")


def test_stage_with_eleven_clusters_is_valid():
    configs = [{"label": "Etapa especial", "start": 1, "end": 1, "clusters": 11, "spf": 4}]
    data = make_source_workbook(stage_count=1, configs=configs)

    result = analyze_workbook(data, "pozo.xlsx")

    assert len(result.stages) == 1
    assert len(result.clusters) == 11


def test_different_number_of_fracture_configurations_is_detected():
    configs = [
        {"label": "A", "start": 1, "end": 2, "clusters": 8, "spf": 4},
        {"label": "B", "start": 3, "end": 4, "clusters": 9, "spf": 4},
        {"label": "C", "start": 5, "end": 6, "clusters": 10, "spf": 5},
        {"label": "D", "start": 7, "end": 8, "clusters": 11, "spf": 5},
    ]
    data = make_source_workbook(stage_count=8, configs=configs)

    result = analyze_workbook(data, "pozo.xlsx")

    assert len(result.fracture_configs) == 4
    assert len(result.stages) == 8


@pytest.mark.parametrize("stage_count", [81, 82])
def test_dynamic_stage_counts_generate_valid_output(stage_count):
    data = make_source_workbook(stage_count=stage_count, survey_rows=12)
    result = analyze_workbook(data, f"LajE-{stage_count}.xlsm")

    generated = generate_finished_workbook(result, TEMPLATE)
    output_workbook = load_workbook(BytesIO(generated.data), data_only=False)
    sheet = output_workbook["Datos terminados"]

    assert len(result.stages) == stage_count
    assert sheet.cell(4, 13).value == 1
    assert sheet.cell(3 + stage_count, 13).value == stage_count
    assert sheet.cell(4, 18).value == "Treatment Interval"
    assert generated.data[:2] == b"PK"


def test_generated_workbook_cleans_residual_template_values(tmp_path):
    dirty_template = tmp_path / "dirty_template.xlsx"
    workbook = load_workbook(TEMPLATE)
    sheet = workbook["Datos terminados"]
    for row in range(12, 18):
        for col in [1, 2, 3, 4, 5, 8, 9, 10, 11, 13, 14, 15, 16, 18, 19, 20]:
            sheet.cell(row, col, "residual")
    workbook.save(dirty_template)

    data = make_source_workbook(stage_count=2, survey_rows=3)
    result = analyze_workbook(data, "pozo.xlsx")
    generated = generate_finished_workbook(result, dirty_template)
    output_workbook = load_workbook(BytesIO(generated.data), data_only=False)
    output_sheet = output_workbook["Datos terminados"]

    for row in range(12, 18):
        for col in [1, 2, 3, 4, 5, 8, 9, 10, 11, 13, 14, 15, 16, 18, 19, 20]:
            assert output_sheet.cell(row, col).value is None


def test_wellbore_has_two_rows_per_stage_in_reverse_order():
    data = make_source_workbook(stage_count=3)
    result = analyze_workbook(data, "pozo.xlsx")
    generated = generate_finished_workbook(result, TEMPLATE)
    sheet = load_workbook(BytesIO(generated.data), data_only=False)["Datos terminados"]

    rows = [sheet.cell(row, 18).value for row in range(4, 4 + result.wellbore_row_count)]

    assert result.wellbore_row_count == 6
    assert rows == [
        "Treatment Interval",
        "Perforations",
        "Treatment Interval",
        "Perforations",
        "Treatment Interval",
        "Perforations",
    ]
    assert sheet.cell(4, 19).value == result.stages[-1].top_md
    assert sheet.cell(8, 19).value == result.stages[0].top_md


def test_output_filename_is_sanitized():
    data = make_source_workbook(well_name="LajE-32(h) PAD / Norte")

    result = analyze_workbook(data, "entrada.xlsm")

    assert result.output_filename == "LajE-32h-PAD-Norte_datos_terminados.xlsx"


def test_generated_xlsx_integrity_and_formula_error_scan():
    data = make_source_workbook(stage_count=4, survey_rows=7)
    generated = process_uploaded_workbook(data, "pozo.xlsm", TEMPLATE)
    workbook = load_workbook(BytesIO(generated.data), data_only=False)
    sheet = workbook["Datos terminados"]
    values = [cell.value for row in sheet.iter_rows() for cell in row]

    assert workbook.sheetnames == ["Datos terminados"]
    assert not {"#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A"} & set(values)
    assert "Archivo .xlsx verificado con openpyxl" in generated.result.checks


def test_generated_numeric_cells_do_not_keep_template_decimal_format():
    data = make_source_workbook(stage_count=3, survey_rows=5)
    generated = process_uploaded_workbook(data, "pozo.xlsm", TEMPLATE)
    workbook = load_workbook(BytesIO(generated.data), data_only=False)
    sheet = workbook["Datos terminados"]

    plain_numeric_cells = [
        "H4",
        "I4",
        "J4",
        "K4",
        "N4",
        "O4",
        "P4",
        "S4",
        "T4",
    ]
    integer_cells = ["B3", "C3", "D3", "E3", "M4"]

    for coordinate in plain_numeric_cells:
        cell = sheet[coordinate]
        assert isinstance(cell.value, (int, float))
        assert cell.number_format == "General"

    for coordinate in integer_cells:
        cell = sheet[coordinate]
        assert isinstance(cell.value, int)
        assert cell.number_format == "0"


def test_duplicate_cluster_numbers_are_rejected():
    data = make_source_workbook(duplicate_cluster_number=True)

    with pytest.raises(InvalidWorkbookError, match="duplicados"):
        analyze_workbook(data, "pozo.xlsx")


def test_tope_must_be_less_than_fondo():
    data = make_source_workbook(bad_depth_stage=1)

    with pytest.raises(InvalidWorkbookError, match="Tope"):
        analyze_workbook(data, "pozo.xlsx")


def test_overrides_are_reported_as_warning():
    data = make_source_workbook(overrides={(1, 1): "Revisar punzado"})

    result = analyze_workbook(data, "pozo.xlsx")

    assert result.override_count == 1
    assert any("sobreescrituras" in warning for warning in result.warnings)


def test_invalid_upload_extension_is_rejected():
    data = make_source_workbook()

    with pytest.raises(InvalidWorkbookError, match=".xlsm o .xlsx"):
        process_uploaded_workbook(data, "pozo.txt", TEMPLATE)


@pytest.mark.parametrize(
    "filename,expected",
    [
        ("LajE-35(h)_version2.xlsm", (82, 784, 225, 3)),
        ("LajE-34(h)_version2.xlsm", (82, 784, 229, 3)),
        ("LajE-33(h)_version2.xlsm", (82, 785, 238, 4)),
        ("LajE-32(h)_version2.xlsm", (81, 779, 612, 4)),
    ],
)
def test_known_wells_if_available(filename, expected):
    path = SAMPLES / filename
    if not path.exists():
        pytest.skip(f"Sample not available: {path}")

    result = analyze_workbook(path.read_bytes(), filename)

    assert (
        len(result.stages),
        len(result.clusters),
        len(result.survey),
        len(result.fracture_configs),
    ) == expected
